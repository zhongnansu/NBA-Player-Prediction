from openpyxl import load_workbook
import xlsxwriter


wb = load_workbook(filename="stats.xlsx")
ws = wb.get_sheet_by_name(u"Seasons_Stats")
data = list(ws.values)
for i in range(0,10000):
    del data[0]
for i in xrange(0,len(data)):
    data[i] = list(data[i])


position = []
for i in data:
    if i[3]  == None:
        position.append(None)
    elif i[3].startswith("C"):
        position.append("C")
    elif i[3].startswith("PG"):
        position.append("PG")
    elif i[3].startswith("SF"):
        position.append("SF")
    elif i[3].startswith("SG"):
        position.append("SG")
    elif i[3].startswith("PF"):
        position.append("PF")
    else:
        anvsinfqke

    for j in iter([30,29,28,27,25,24,23,22,11,9,5,4,3,2,1,0]):
       if j == 24:
           continue
       del i[j]





workbook = xlsxwriter.Workbook("pos.xlsx")
worksheet1 = workbook.add_worksheet()
worksheet2 = workbook.add_worksheet()
worksheet3 = workbook.add_worksheet()
worksheet4 = workbook.add_worksheet()
worksheet5 = workbook.add_worksheet()

row1, row2, row3, row4, row5 = 0,0,0,0,0
for i in range(0, len(data)):
    if position[i] == "C":
        for j in range(0, len(data[i])):
            worksheet1.write(row1, j, data[i][j])
        row1 = row1 + 1
    elif position[i] == "PG":
        for j in range(0, len(data[i])):
            worksheet2.write(row2, j, data[i][j])
        row2 = row2 + 1
    elif position[i] == "SF":
        for j in range(0, len(data[i])):
            worksheet3.write(row3, j, data[i][j])
        row3 = row3 + 1
    elif position[i] == "SG":
        for j in range(0, len(data[i])):
            worksheet4.write(row4, j, data[i][j])
        row4 = row4 + 1
    elif position[i] == "PF":
        for j in range(0, len(data[i])):
            worksheet5.write(row5, j, data[i][j])
        row5 = row5 + 1

workbook.close()